from __future__ import annotations

import json
import wave
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image

import app as space_app


class _DumpableResponse:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def model_dump(self, mode: str = "json") -> dict[str, Any]:
        assert mode == "json"
        return self.payload


def test_space_health_reports_strict_local_runtime() -> None:
    summary, payload = space_app.space_health()

    assert "Privacy: strict" in summary
    assert "No external model APIs" in summary
    assert "Ready required models:" in summary
    assert "Ready required models: asr, text, vision" in summary
    assert "Pending required models: none" in summary
    assert payload["local_only"] is True
    assert payload["privacy_mode"] == "strict"
    assert payload["allow_web"] is False
    assert payload["telemetry_enabled"] is False


def test_first_screen_disclosure_names_space_and_local_runtime() -> None:
    disclosure = space_app.first_screen_disclosure()

    assert "Hugging Face Space compute" in disclosure
    assert "external model APIs" in disclosure
    assert "cloud OCR" in disclosure
    assert "remote telemetry" in disclosure
    assert "Laptop-local mode" in disclosure


def test_demo_first_screen_contains_privacy_disclosure() -> None:
    markdown_blocks = [
        block for block in space_app.demo.blocks.values() if type(block).__name__ == "Markdown"
    ]
    values = "\n".join(str(getattr(block, "value", "")) for block in markdown_blocks)

    assert "# local-lm" in values
    assert "Hosted Space note" in values
    assert "Hugging Face Space compute" in values
    assert "does not call external model APIs" in values


def test_advanced_runtime_panel_is_closed_by_default() -> None:
    accordions = [
        block for block in space_app.demo.blocks.values() if type(block).__name__ == "Accordion"
    ]

    runtime_panel = next(
        block
        for block in accordions
        if getattr(block, "label", "") == space_app.ADVANCED_RUNTIME_LABEL
    )
    assert runtime_panel.open is False


def test_summarize_url_is_blocked_by_default() -> None:
    summary, payload = space_app.summarize_article("", "https://example.com/article", False)

    assert payload["task"] == "summarize_url"
    assert payload["status"] == "blocked"
    assert payload["local_only"] is True
    assert "Web access is disabled" in summary


def test_summarize_article_shows_sensitive_output_warning() -> None:
    summary, payload = space_app.summarize_article(
        "This article explains a medical treatment and legal notice.",
        "",
        False,
    )

    assert payload["task"] == "summarize_wikipedia"
    assert payload["human_review_required"] is True
    assert "qualified human" in summary
    assert "warnings:" in summary.casefold()


def test_ask_assistant_returns_local_unavailable_warning() -> None:
    answer, payload = space_app.ask_assistant("How do I save this invoice?")

    assert payload["task"] == "general_local_assistant"
    assert payload["status"] == "stub"
    assert payload["local_only"] is True
    assert "local text model is unavailable" in answer.casefold()
    assert "no remote assistant service" in answer.casefold()
    assert "warnings:" in answer.casefold()


def test_ask_assistant_shows_sensitive_output_warning() -> None:
    answer, payload = space_app.ask_assistant("Can you explain this medical bill?")

    assert payload["task"] == "general_local_assistant"
    assert payload["human_review_required"] is True
    assert "qualified human" in answer
    assert "warnings:" in answer.casefold()


def test_ask_assistant_requires_question() -> None:
    answer, payload = space_app.ask_assistant("   ")

    assert answer == "Type a question first."
    assert payload == {}


def test_text_result_export_writes_txt_xlsx_and_pdf() -> None:
    payload = {
        "task": "general_local_assistant",
        "status": "stub",
        "warnings": ["Local text model is unavailable."],
        "human_review_required": True,
        "result": {"source": "local_fallback"},
    }

    files, status = space_app.save_text_result_exports(
        "unit test answer",
        "Simple answer.\n\nWarnings:\nLocal text model is unavailable.",
        payload,
        ["TXT", "XLSX", "PDF"],
    )

    paths = [Path(path) for path in files]
    assert status == "Saved 3 local file(s)."
    assert {path.suffix for path in paths} == {".txt", ".xlsx", ".pdf"}
    assert all(path.exists() for path in paths)
    txt_path = next(path for path in paths if path.suffix == ".txt")
    assert "Simple answer." in txt_path.read_text(encoding="utf-8")
    pdf_path = next(path for path in paths if path.suffix == ".pdf")
    assert pdf_path.read_bytes().startswith(b"%PDF")
    xlsx_path = next(path for path in paths if path.suffix == ".xlsx")
    workbook = load_workbook(xlsx_path)
    assert "metadata" in workbook.sheetnames
    metadata_rows = {
        str(row[0].value): row[1].value for row in workbook["metadata"].iter_rows(min_row=2)
    }
    assert metadata_rows["task"] == "general_local_assistant"
    assert metadata_rows["human_review_required"] is True


def test_text_result_export_rejects_empty_result() -> None:
    files, status = space_app.save_text_result_exports("empty", "   ", {}, ["TXT"])

    assert files == []
    assert status == "Nothing to save yet."


def test_document_conversion_writes_requested_local_exports(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    document = tmp_path / "invoice.pdf"
    document.write_text("synthetic invoice", encoding="utf-8")

    monkeypatch.setattr(
        space_app,
        "gateway_document_to_excel",
        lambda request: _DumpableResponse(
            {
                "task": "document_to_excel",
                "status": "stub",
                "local_only": True,
                "confidence": 0.0,
                "human_review_required": True,
                "warnings": ["Local document model is unavailable; review required."],
                "result": {
                    "rows": [],
                    "schema": {"human_review_required": True},
                },
            }
        ),
    )

    payload_text, files, warnings = space_app.convert_document(
        str(document), ["JSON", "XLSX", "TXT", "PDF"]
    )

    payload = json.loads(payload_text)
    assert payload["task"] == "document_to_excel"
    assert payload["local_only"] is True
    assert payload["human_review_required"] is True
    assert payload["confidence"] == 0.0
    assert payload["result"]["schema"]["human_review_required"] is True
    assert warnings
    assert {Path(path).suffix for path in files} == {".json", ".xlsx", ".txt", ".pdf"}
    assert all(Path(path).exists() for path in files)
    pdf_path = next(Path(path) for path in files if Path(path).suffix == ".pdf")
    assert pdf_path.read_bytes().startswith(b"%PDF")


def test_document_conversion_extracts_demo_invoice_sample() -> None:
    invoice_path = Path("samples/demo/invoice_sample.txt")

    payload_text, files, warnings = space_app.convert_document(str(invoice_path), ["JSON", "XLSX"])

    payload = json.loads(payload_text)
    assert payload["status"] == "ok"
    assert payload["confidence"] == 0.85
    assert payload["result"]["schema"]["document_type"] == "invoice"
    assert payload["result"]["schema"]["total"] == 1770.0
    assert payload["result"]["schema"]["tax_amount"] == 270.0
    assert payload["human_review_required"] is True
    assert warnings
    assert {Path(path).suffix for path in files} == {".json", ".xlsx"}


def test_document_xlsx_export_includes_review_metadata() -> None:
    invoice_path = Path("samples/demo/invoice_sample.txt")

    payload_text, files, warnings = space_app.convert_document(str(invoice_path), ["XLSX"])

    payload = json.loads(payload_text)
    xlsx_path = next(Path(path) for path in files if Path(path).suffix == ".xlsx")
    workbook = load_workbook(xlsx_path)
    assert "metadata" in workbook.sheetnames
    metadata_rows = {
        str(row[0].value): row[1].value for row in workbook["metadata"].iter_rows(min_row=2)
    }
    assert metadata_rows["task"] == "document_to_excel"
    assert metadata_rows["status"] == payload["status"]
    assert float(metadata_rows["confidence"]) == payload["confidence"]
    assert metadata_rows["human_review_required"] is True
    assert "Synthetic/local text extraction" in metadata_rows["warnings"]
    assert warnings


def test_image_handlers_return_local_stub_outputs(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    image_path = tmp_path / "photo.png"
    Image.new("RGB", (16, 16), color="white").save(image_path)

    monkeypatch.setattr(
        space_app,
        "gateway_describe_image",
        lambda request: _DumpableResponse(
            {
                "task": "describe_image",
                "status": "stub",
                "local_only": True,
                "human_review_required": True,
                "warnings": [
                    "Local vision model is unavailable; selected file stayed local.",
                    "No cloud OCR or remote inference was used.",
                ],
                "result": {
                    "description": (
                        "Local vision model is unavailable. The selected image stayed local; "
                        "no cloud OCR or remote inference was used."
                    ),
                    "schema": {"human_review_required": True},
                },
            }
        ),
    )
    monkeypatch.setattr(
        space_app,
        "gateway_translate_image_text",
        lambda request: _DumpableResponse(
            {
                "task": "translate_image_text",
                "status": "stub",
                "local_only": True,
                "human_review_required": True,
                "warnings": [
                    "Local OCR/translation model is unavailable; selected file stayed local.",
                    "No cloud OCR or remote inference was used.",
                ],
                "result": {
                    "text": (
                        "Local OCR/translation model is unavailable. The selected file "
                        "stayed local; no cloud OCR or remote inference was used."
                    ),
                    "schema": {"human_review_required": True},
                },
            }
        ),
    )

    description, describe_payload = space_app.describe_photo(str(image_path))
    translation, translate_payload = space_app.translate_visible_text(str(image_path), "Hindi")

    assert describe_payload["task"] == "describe_image"
    assert translate_payload["task"] == "translate_image_text"
    assert describe_payload["human_review_required"] is True
    assert translate_payload["human_review_required"] is True
    assert "schema" in describe_payload["result"]
    assert "schema" in translate_payload["result"]
    assert "local vision model is unavailable" in description.casefold()
    assert "cloud ocr" in description.casefold()
    assert "local ocr/translation model is unavailable" in translation.casefold()
    assert "cloud ocr" in translation.casefold()
    assert "warnings:" in description.casefold()
    assert "warnings:" in translation.casefold()


def test_speech_handler_returns_local_stub_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    audio_path = tmp_path / "voice.wav"
    with wave.open(str(audio_path), "wb") as wav_file:
        wav_file.setnchannels(1)
        wav_file.setsampwidth(2)
        wav_file.setframerate(16_000)
        wav_file.writeframes(b"\x00\x00" * 160)

    monkeypatch.setattr(
        space_app,
        "gateway_speech_to_text",
        lambda request: _DumpableResponse(
            {
                "task": "speech_to_text",
                "status": "stub",
                "local_only": True,
                "human_review_required": False,
                "warnings": [
                    "Local ASR service is unavailable.",
                    "No remote speech service was called.",
                ],
                "result": {
                    "text": (
                        "Local ASR service is unavailable. "
                        "No remote speech service was called."
                    ),
                    "asr_endpoint": "http://127.0.0.1:8090",
                    "model_ready": False,
                },
            }
        ),
    )

    transcript, payload = space_app.transcribe_speech(str(audio_path))

    assert payload["task"] == "speech_to_text"
    assert payload["local_only"] is True
    assert "local asr service is unavailable" in transcript.casefold()
    assert "warnings:" in transcript.casefold()
    assert "no remote speech service" in transcript.casefold()


def test_speech_handler_marks_indian_non_english_asr_experimental(tmp_path: Path) -> None:
    audio_path = tmp_path / "voice.wav"
    with wave.open(str(audio_path), "wb") as wav_file:
        wav_file.setnchannels(1)
        wav_file.setsampwidth(2)
        wav_file.setframerate(16_000)
        wav_file.writeframes(b"\x00\x00" * 160)

    transcript, payload = space_app.transcribe_speech(str(audio_path), "hi", "india", False)

    assert payload["task"] == "speech_to_text"
    assert payload["status"] == "stub"
    assert payload["result"]["language"] == "hi"
    assert payload["result"]["region"] == "india"
    assert payload["result"]["experimental"] is True
    assert payload["result"]["unsupported_language"] is True
    assert payload["human_review_required"] is True
    assert "warnings:" in transcript.casefold()
    assert any("experimental ASR" in warning for warning in payload["warnings"])


def test_router_handler_keeps_web_disabled_without_opt_in(tmp_path: Path) -> None:
    document = tmp_path / "bill.pdf"
    document.write_text("synthetic bill", encoding="utf-8")

    decision = space_app.route_intent(
        "convert this to Excel",
        str(document),
        "",
        False,
        space_app.AUTO_ROUTE,
    )

    assert decision["task"] == "document_to_excel"
    assert decision["provider"] == "local"
    assert decision["privacy_mode"] == "strict"
    assert decision["allow_web"] is False
    assert decision["manual_override"] is False


def test_manual_route_override_selects_requested_local_task() -> None:
    decision = space_app.route_intent(
        "please do whatever is best",
        None,
        "",
        False,
        "speech_to_text",
    )

    assert decision["task"] == "speech_to_text"
    assert decision["provider"] == "local"
    assert decision["model_key"] == "asr"
    assert decision["allow_web"] is False
    assert decision["manual_override"] is True


def test_demo_samples_are_available_and_loadable() -> None:
    samples = space_app.demo_sample_paths()

    assert {"question", "article", "invoice", "bank_statement"} <= set(samples)
    assert "private files" in space_app.load_demo_question().casefold()
    assert "local-first assistive technology" in space_app.load_demo_article().casefold()
    assert space_app.load_demo_document("invoice")
    assert space_app.load_demo_document("bank statement")


def test_generated_demo_media_are_available_and_local() -> None:
    image_path = Path(space_app.load_demo_image())
    audio_path = Path(space_app.load_demo_audio())

    assert image_path.exists()
    assert image_path.suffix == ".png"
    assert audio_path.exists()
    assert audio_path.suffix == ".wav"
    Image.open(image_path).verify()
    with wave.open(str(audio_path), "rb") as wav_file:
        assert wav_file.getnchannels() == 1
        assert wav_file.getframerate() == 16_000


def test_privacy_disclosure_and_model_budget_match_hackathon_constraints() -> None:
    disclosure = space_app.privacy_disclosure()
    budget = space_app.model_budget_status()

    assert "Hugging Face Space compute" in disclosure
    assert "No external model APIs" in disclosure
    assert budget["within_cap"] is True
    assert budget["parameter_cap_b"] == 32
    assert budget["privacy_mode"] == "strict"
    assert "text" in budget["ready_required_models"]
    assert "vision" in budget["ready_required_models"]
    assert "text" in budget["artifact_present_required_models"]
    assert "vision" in budget["checksum_configured_required_models"]
    assert budget["required_model_readiness"]["text"]["ready"] is True


def test_demo_readiness_distinguishes_demo_path_from_release_gate() -> None:
    summary, payload = space_app.demo_readiness_status()

    assert "Demo Readiness" in summary
    assert payload["demo_ready"] is True
    assert payload["release_ready"] is True
    assert payload["checks"]["strict_privacy"] is True
    assert payload["checks"]["model_budget_under_32b"] is True
    assert payload["checks"]["text_model_checksum_ready"] is True
    assert payload["checks"]["vision_model_checksum_ready"] is True
    assert payload["checks"]["asr_model_checksum_ready"] is True
    assert payload["checks"]["text_model_artifact_present"] is True
    assert payload["checks"]["vision_model_artifact_present"] is True
    assert payload["checks"]["asr_model_artifact_present"] is True
    assert payload["checks"]["text_model_ready"] is True
    assert payload["checks"]["vision_model_ready"] is True
    assert payload["checks"]["asr_model_ready"] is True
    assert payload["checks"]["release_gate_metadata_ready"] is True
    assert payload["release_gate"]["status"] == "ok"
    assert payload["release_gate"]["checksum_verification"] is False
    assert payload["checks"]["demo_samples_available"] is True
    assert "question" in payload["sample_paths"]
    assert payload["checks"]["generated_demo_media_available"] is True
    assert payload["blocking_release_gaps"] == []
    assert any("--mock-model-endpoints" in command for command in payload["next_commands"])
    assert not any("--write-manifest-checksum" in command for command in payload["next_commands"])
    assert any("--require-real-model-services" in command for command in payload["next_commands"])
    assert "Release gate metadata: yes" in summary


def test_demo_readiness_rejects_checksum_without_local_artifact(
    monkeypatch,
    tmp_path: Path,
) -> None:
    image_path = tmp_path / "demo.png"
    audio_path = tmp_path / "demo.wav"
    image_path.write_bytes(b"fake image placeholder")
    audio_path.write_bytes(b"fake audio placeholder")

    monkeypatch.setattr(space_app, "load_demo_image", lambda: str(image_path))
    monkeypatch.setattr(space_app, "load_demo_audio", lambda: str(audio_path))
    monkeypatch.setattr(
        space_app,
        "space_health",
        lambda: (
            "Status: ok",
            {
                "privacy_mode": "strict",
                "allow_web": False,
                "telemetry_enabled": False,
            },
        ),
    )
    monkeypatch.setattr(
        space_app,
        "model_budget_status",
        lambda: {
            "within_cap": True,
            "ready_required_models": ["vision"],
            "pending_required_models": ["text", "asr"],
            "checksum_configured_required_models": ["text", "vision"],
            "artifact_present_required_models": ["vision"],
            "local_only": True,
            "privacy_mode": "strict",
        },
    )

    summary, payload = space_app.demo_readiness_status()

    assert "needs attention" in summary
    assert payload["demo_ready"] is False
    assert payload["release_ready"] is False
    assert payload["checks"]["text_model_checksum_ready"] is True
    assert payload["checks"]["text_model_artifact_present"] is False
    assert payload["checks"]["text_model_ready"] is False
    assert "Required text model artifact/checksum is not ready." in payload["blocking_release_gaps"]


def test_local_demo_flow_runs_all_sample_tasks_without_external_apis() -> None:
    summary, payload = space_app.run_local_demo_flow()

    assert "Local Demo Flow" in summary
    assert payload["status"] == "ok"
    assert payload["local_only"] is True
    assert payload["external_model_apis"] is False
    assert payload["cloud_ocr"] is False
    assert payload["remote_telemetry"] is False
    assert payload["checks"] == {
        "samples_loaded": True,
        "generated_media_loaded": True,
        "document_exports_created": True,
        "all_steps_local": True,
        "structured_outputs": True,
        "visible_warnings": True,
    }
    assert set(payload["steps"]) == {"ask", "read", "document", "image", "speech"}
    assert payload["steps"]["ask"]["task"] == "general_local_assistant"
    assert payload["steps"]["read"]["task"] == "summarize_wikipedia"
    assert payload["steps"]["document"]["task"] == "document_to_excel"
    assert payload["steps"]["image"]["task"] == "describe_image"
    assert payload["steps"]["speech"]["task"] == "speech_to_text"
    assert all(Path(path).exists() for path in payload["steps"]["document"]["downloads"])


def test_settings_tab_exposes_local_demo_flow_button() -> None:
    labels_or_values = [
        getattr(block, "label", "") or getattr(block, "value", "")
        for block in space_app.demo.blocks.values()
    ]

    assert "Run Local Demo Samples" in labels_or_values
    assert "Local demo flow" in labels_or_values


def test_accessibility_settings_emit_high_contrast_stylesheet() -> None:
    enabled_status, enabled_style = space_app.accessibility_settings(True)
    disabled_status, disabled_style = space_app.accessibility_settings(False)

    assert enabled_status == "High contrast mode is on."
    assert disabled_status == "High contrast mode is off."
    assert "local-lm-accessibility-style" in enabled_style
    assert "background: #000" in enabled_style
    assert "button.primary" in enabled_style
    assert disabled_style == space_app.DEFAULT_ACCESSIBILITY_CSS


def test_base_css_uses_large_controls_for_all_buttons() -> None:
    assert ".gradio-container button" in space_app.APP_CSS
    assert "min-height: 52px" in space_app.APP_CSS
    assert "font-size: 18px !important" in space_app.APP_CSS
    assert "line-height: 1.5" in space_app.APP_CSS


def test_demo_embeds_base_accessibility_stylesheet() -> None:
    html_blocks = [
        block for block in space_app.demo.blocks.values() if type(block).__name__ == "HTML"
    ]
    styles = "\n".join(str(getattr(block, "value", "")) for block in html_blocks)

    assert "local-lm-base-style" in styles
    assert ".gradio-container button" in styles
    assert "min-height: 52px" in styles


def test_launch_helper_uses_imported_demo_without_separate_css(monkeypatch) -> None:
    captured: dict[str, object] = {}

    def fake_launch(**kwargs):
        captured["kwargs"] = kwargs

    monkeypatch.setattr(space_app.demo, "launch", fake_launch)

    space_app.launch()

    assert captured["kwargs"] == {}


def test_text_oriented_tabs_expose_save_controls() -> None:
    checkbox_groups = [
        block for block in space_app.demo.blocks.values() if type(block).__name__ == "CheckboxGroup"
    ]
    save_groups = [
        block for block in checkbox_groups if getattr(block, "label", "") == "Save result as"
    ]
    labels_or_values = [
        getattr(block, "label", "") or getattr(block, "value", "")
        for block in space_app.demo.blocks.values()
    ]

    assert len(save_groups) == 4
    assert "Save Answer" in labels_or_values
    assert "Save Summary" in labels_or_values
    assert "Save Image Result" in labels_or_values
    assert "Save Transcript" in labels_or_values


def test_speech_tab_exposes_asr_language_region_controls() -> None:
    labels_or_values = [
        getattr(block, "label", "") or getattr(block, "value", "")
        for block in space_app.demo.blocks.values()
    ]

    assert "Language code" in labels_or_values
    assert "Region" in labels_or_values
    assert "Allow experimental ASR" in labels_or_values


def test_space_requirements_stay_runtime_only() -> None:
    requirements = Path("requirements.txt").read_text(encoding="utf-8").casefold()

    assert "httpx2" not in requirements
    for training_package in (
        "transformers",
        "datasets",
        "accelerate",
        "peft",
        "trl",
        "bitsandbytes",
    ):
        assert training_package not in requirements
